import openpyxl

files = "..\\testdata\\testdata.xlsx"


class Excel:
    def read_data(self, sheet, cell_chords):
        self.workbook = openpyxl.load_workbook(files)
        self.sheet = self.workbook[sheet]
        return self.sheet[cell_chords].value

    def write_data(self, sheet, cell_chords, data):
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        sheet[cell_chords].value = data
        workbook.save(files)

    def row_count(self, sheet):
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        rows = sheet.max_row
        return rows

    def col_count(self, sheet):
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        cols = sheet.max_column
        return cols
